import re, json, html, time, os
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

URL = "https://pickleball.global/index.php?option=com_globalpicranking&task=gentelmens.singles"
HEAD = {"X-Requested-With": "XMLHttpRequest",
        "Content-Type": "application/x-www-form-urlencoded",
        "User-Agent": "Mozilla/5.0"}

YEARS = [2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
AGES = ["19", "35", "50", "60", "70", "75", "80"]

COLS = ["#", "PLAYER", "AGE", "GPR POINTS", "EVENTS PLAYED", "EVENT TYPE",
        "START DATE", "END DATE", "MATCHES (W-L)", "GAMES (W-L)", "POINTS (W-L)",
        "AGE GROUP", "IFP RATING", "COUNTRY", "ACTION"]

def fetch(age, year):
    data = {
        "type": "Mens Singles", "ageGroup": age, "country": "", "state": "",
        "gender": "", "pageFor": "GENTLEMEN", "continent": "", "gprYear": str(year),
        "includePro": "1", "nationalTournaments": "", "race": "", "wpceu": "", "tids": "",
    }
    for attempt in range(4):
        try:
            r = requests.post(URL, headers=HEAD, data=data, timeout=180)
            raw = r.text
            i = raw.find('{"data"')
            if i < 0:
                return []
            d = json.loads(raw[i:])
            return parse_rows(d["data"])
        except Exception as e:
            if attempt == 3:
                print(f"    FAIL age {age} year {year}: {e}")
                return []
            time.sleep(2 ** attempt)

def parse_rows(table_html):
    trs = re.findall(r'<tr>(.*?)</tr>', table_html, re.S)
    out = []
    for tr in trs[1:]:
        cells = [html.unescape(re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', x)).strip())
                 for x in re.findall(r'<td[^>]*>(.*?)</td>', tr, re.S)]
        if len(cells) >= 14:
            out.append(cells[:15] if len(cells) >= 15 else cells + [""] * (15 - len(cells)))
    return out

def is_india(row):
    return row[13].strip().lower() == "india"

def main():
    # data[year][age] = list of india rows
    alldata = {}
    for year in YEARS:
        alldata[year] = {}
        for age in AGES:
            rows = fetch(age, year)
            ind = [r for r in rows if is_india(r)]
            alldata[year][age] = ind
            print(f"year {year} age {age:>2}+ : total={len(rows):5d} india={len(ind):4d}")
            time.sleep(0.4)

    wb = Workbook()
    summ = wb.active
    summ.title = "Summary"
    summ.append(["Pickleball Global (GPR) Rankings — INDIA — by Year & Age Group"])
    summ.append([])
    summ.append(["Year"] + [f"{a}+" for a in AGES] + ["Total"])
    for year in YEARS:
        counts = [len(alldata[year][a]) for a in AGES]
        summ.append([year] + counts + [sum(counts)])
    summ["A1"].font = Font(bold=True, size=12)
    for c in range(1, len(AGES) + 3):
        summ.cell(row=3, column=c).font = Font(bold=True)
    summ.column_dimensions["A"].width = 10

    green = PatternFill("solid", fgColor="1B5E20")
    # One tab per age group, all years stacked with a YEAR column
    OUT_COLS = ["YEAR", "RANK"] + COLS[1:]   # drop the original '#' label dup
    for age in AGES:
        ws = wb.create_sheet(title=f"{age}+ India")
        ws.append([f"GPR Rankings — India — {age}+ — all years 2019-2026"])
        ws.append([])
        ws.append(OUT_COLS)
        total = 0
        for year in YEARS:
            for r in alldata[year][age]:
                ws.append([year] + r)   # r[0] is rank '#'
                total += 1
        if total == 0:
            ws.append(["(no India players in this age group)"])
        ws["A1"].font = Font(bold=True, size=12)
        for c in range(1, len(OUT_COLS) + 1):
            cell = ws.cell(row=3, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = green
        widths = [7, 7, 26, 6, 11, 9, 28, 12, 12, 13, 13, 13, 11, 11, 12, 9]
        for i, w in enumerate(widths[:len(OUT_COLS)], 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
        ws.freeze_panes = "A4"

    # Combined "All India" sheet
    ws = wb.create_sheet(title="All India", index=1)
    ws.append(["AGE GROUP", "YEAR", "RANK"] + COLS[1:])
    for c in range(1, 4 + len(COLS[1:])):
        ws.cell(row=1, column=c).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=c).fill = green
    ws.freeze_panes = "A2"
    grand = 0
    for age in AGES:
        for year in YEARS:
            for r in alldata[year][age]:
                ws.append([f"{age}+", year] + r)
                grand += 1

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "pickleball_global_india_rankings.xlsx")
    wb.save(out)
    print(f"\nSAVED {out}")
    print(f"Total India ranking rows (all years/ages): {grand}")

if __name__ == "__main__":
    main()
