"""
Scrape West Zone Pickleball Championship 2026 match results from the public
Khel Club portal and export to Excel (one tab per category) + CSVs.

The championship (public short link https://app.khelclub.co/233df) runs
29-31 May 2026 and is a single Khel Club tournament (6a1696258c90b13d8df233df)
containing 24 categories:
  * 15 youth categories  (U-12 / U-14 / U-18)         -> mostly day 1
  * 9 adult categories   (Open / 35+ / Beginner-Int.) -> days 2-3

Each category's fixtures page server-renders its COMPLETE match list (all pools
and all knockout rounds, across every day) into the Next.js RSC flight payload,
so one fetch per category yields that category's full multi-day data. The
`dateTab` query param is only a UI hint and does not change the payload.

The category display name is read from the embedded `initialCategory.name`
field, so names do not need to be hard-coded.
"""
import re, json, time, os, csv
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE = "https://app.khelclub.co"
ORG = "6a1694b88c90b13d8df231c7"
TOUR = "6a1696258c90b13d8df233df"
DATE = "2026-05-31"   # UI hint only; payload carries all days regardless
DEC = json.JSONDecoder()

# Category IDs (names auto-detected from each payload's initialCategory.name).
# Youth ids are sequential (e2..f0); adult ids collected from the public landing.
CAT_IDS = [
    # ---- Youth (U-12 / U-14 / U-18) ----
    "6a1696258c90b13d8df233e2", "6a1696258c90b13d8df233e3", "6a1696258c90b13d8df233e4",
    "6a1696258c90b13d8df233e5", "6a1696258c90b13d8df233e6", "6a1696258c90b13d8df233e7",
    "6a1696258c90b13d8df233e8", "6a1696258c90b13d8df233e9", "6a1696258c90b13d8df233ea",
    "6a1696258c90b13d8df233eb", "6a1696258c90b13d8df233ec", "6a1696258c90b13d8df233ed",
    "6a1696258c90b13d8df233ee", "6a1696258c90b13d8df233ef", "6a1696258c90b13d8df233f0",
    # ---- Adult (Open / 35+ / Beginner-Intermediate) ----
    "6a16b2118c90b13d8df2807c", "6a16b2188c90b13d8df28099", "6a16b2201770aec6bcf37e09",
    "6a16bc074706ff44e4b9e278", "6a16bc154706ff44e4b9e290", "6a16d0cd8c90b13d8df312ce",
    "6a16d0d58c90b13d8df312e2", "6a16d0da2f52f8f4470d53be", "6a199b524828cee159b98d31",
]

def fetch_flight(cat_id):
    url = f"{BASE}/organizer-profile/{ORG}/tournament/{TOUR}/category/{cat_id}/fixtures?dateTab=date-{DATE}"
    html = requests.get(url, timeout=30).text
    parts = []
    for p in re.findall(r'self\.__next_f\.push\(\[\d+,(.*?)\]\)</script>', html, re.S):
        try: parts.append(json.loads(p))
        except Exception: pass
    return "".join(x for x in parts if isinstance(x, str))

def objects_with_key(blob, key, maxspan=30000):
    """Extract every self-contained JSON object that directly contains `key`."""
    out, pos = [], 0
    while True:
        k = blob.find('"%s"' % key, pos)
        if k < 0: break
        j, hit = k, False
        while True:
            j = blob.rfind('{', 0, j)
            if j < 0 or k - j > maxspan: break
            try:
                o, end = DEC.raw_decode(blob, j)
                if isinstance(o, dict) and key in o and end > k:
                    out.append(o); pos = end; hit = True; break
            except Exception:
                pass
        if not hit: pos = k + 1
    return out

def category_name(blob):
    """Read the category display name from the embedded initialCategory object."""
    m = re.search(r'"initialCategory":\{"id":"[a-f0-9]{24}","name":"([^"]+)"', blob)
    if m: return m.group(1)
    m = re.search(r'"name":"([^"]+)"[^{}]{0,80}?"matchStructure"', blob)
    return m.group(1) if m else "Unknown Category"

def label(team, tmap):
    if not team or not isinstance(team, dict): return "TBD"
    tid = team.get('id') or team.get('_id')
    if tid and tmap.get(tid): return tmap[tid]
    return team.get('name', 'TBD')

def score_str(scores):
    if not scores: return ""
    if all((s.get('team1') or 0) == 0 and (s.get('team2') or 0) == 0 for s in scores):
        return ""  # not yet played
    return ", ".join(f"{s.get('team1')}-{s.get('team2')}"
                     for s in sorted(scores, key=lambda x: x.get('set', 0)))

def containers(blob):
    """Extract every self-contained object holding a `matches` list
    (a round-robin pool or a knockout/elimination round)."""
    out, pos = [], 0
    while True:
        k = blob.find('"matches"', pos)
        if k < 0: break
        j, hit = k, False
        while True:
            j = blob.rfind('{', 0, j)
            if j < 0 or k - j > 80000: break
            try:
                o, end = DEC.raw_decode(blob, j)
                if isinstance(o, dict) and isinstance(o.get('matches'), list) and end > k:
                    out.append(o); pos = end; hit = True; break
            except Exception:
                pass
        if not hit: pos = k + 1
    return out

def stage_label(c):
    nm = c.get('name', '') or ''
    if c.get('roundTitle') == 'Elimination' or 'Final' in nm or 'Quarter' in nm or 'Semi' in nm:
        return nm or 'Knockout'
    if nm.lower().startswith('pool'):
        return f"Round Robin ({nm})"
    return "Round Robin"

def stage_order(c):
    if c.get('roundTitle') == 'Elimination':
        return (1, c.get('roundNumber', 99))
    return (0, c.get('name', ''))

def parse_category(blob):
    name = category_name(blob)
    structure = (re.findall(r'"matchStructure":"(\w+)"', blob) or [""])[0]
    tmap = {}
    for t in objects_with_key(blob, 'players'):
        tid = t.get('id') or t.get('_id')
        names = [p.get('name', '') for p in t.get('players', []) if isinstance(p, dict) and p.get('name')]
        if tid and names:
            tmap[tid] = " & ".join(names)
    rows, seen = [], set()
    for c in sorted(containers(blob), key=stage_order):
        stage = stage_label(c)
        for m in c['matches']:
            if not isinstance(m, dict): continue
            mid = m.get('id') or m.get('_id')
            if mid and mid in seen: continue
            if mid: seen.add(mid)
            t1 = label(m.get('team1'), tmap)
            t2 = label(m.get('team2'), tmap)
            sc = score_str(m.get('scores') or [])
            win = label(m.get('winnerTeam'), tmap) if m.get('winnerTeam') else ""
            st = m.get('status', '') or ("SCHEDULED" if not sc else "")
            flags = []
            if m.get('isWalkover'): flags.append('Walkover')
            if m.get('isBye'): flags.append('Bye')
            if m.get('isForfeit'): flags.append('Forfeit')
            if t1 == "TBD" and t2 == "TBD" and not sc:
                continue
            rows.append([stage, t1, t2, sc, win, st, ", ".join(flags)])
    return name, structure, rows

HEADERS = ["Stage", "Team 1", "Team 2", "Score", "Winner", "Status", "Notes"]
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def safe_title(name, used):
    t = re.sub(r"[\[\]\:\*\?/\\']", "", name)[:31] or "Category"
    base, i = t, 2
    while t in used:
        t = f"{base[:28]}_{i}"; i += 1
    used.add(t); return t

def csv_name(name):
    return re.sub(r"[ /']+", "_", name).strip("_") + ".csv"

def main():
    data = []   # (name, structure, rows)
    for cid in CAT_IDS:
        try:
            blob = fetch_flight(cid)
            name, structure, rows = parse_category(blob)
            data.append((name, structure, rows))
            comp = sum(1 for r in rows if str(r[5]).upper() == "COMPLETED")
            print(f"{name:38s} {structure:18s} matches={len(rows):3d} completed={comp}")
        except Exception as e:
            print(f"{cid} ERROR {e}")
        time.sleep(0.3)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["West Zone Pickleball Championship 2026 — Musclebar Sports Club, Pune — 29-31 May 2026"])
    summary.append([])
    summary.append(["Category", "Format", "Matches", "Completed", "Remaining"])
    for name, structure, rows in data:
        comp = sum(1 for r in rows if str(r[5]).upper() == "COMPLETED")
        summary.append([name, structure, len(rows), comp, len(rows) - comp])
    summary["A1"].font = Font(bold=True, size=12)
    for c in range(1, 6):
        summary.cell(row=3, column=c).font = Font(bold=True)
    for col, w in zip("ABCDE", [38, 20, 9, 11, 11]):
        summary.column_dimensions[col].width = w

    cdir = os.path.join(REPO, "match_results_csv")
    os.makedirs(cdir, exist_ok=True)
    used = {"Summary"}
    combined = [["Category", "Format"] + HEADERS]
    for name, structure, rows in data:
        ws = wb.create_sheet(title=safe_title(name, used))
        ws.append([f"{name}  |  {structure}  |  29-31 May 2026"])
        ws.append([])
        ws.append(HEADERS)
        for r in rows: ws.append(r)
        if not rows: ws.append(["(no match data available yet)"])
        ws["A1"].font = Font(bold=True, size=12)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=3, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E7D32")
        for col, w in zip("ABCDEFG", [20, 26, 26, 16, 26, 13, 14]):
            ws.column_dimensions[col].width = w
        with open(os.path.join(cdir, csv_name(name)), "w", newline="") as f:
            wr = csv.writer(f); wr.writerow(HEADERS); wr.writerows(rows)
        for r in rows:
            combined.append([name, structure] + r)

    with open(os.path.join(cdir, "ALL_categories_combined.csv"), "w", newline="") as f:
        csv.writer(f).writerows(combined)

    out = os.path.join(REPO, "west_zone_pickleball_2026_results.xlsx")
    wb.save(out)
    total = len(combined) - 1
    done = sum(1 for r in combined[1:] if str(r[8]).upper() == "COMPLETED")
    print(f"\nSAVED {out}")
    print(f"Categories: {len(data)} | Matches: {total} | Completed: {done} | Remaining: {total - done}")

if __name__ == "__main__":
    main()
